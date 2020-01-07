#!/usr/bin/python3
from django.shortcuts import render, render_to_response
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse as HR
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.db.models import Count, Sum, Avg, Max, Min

from notebook.models import User
from icu.models import QualityControl, ApacheII

from pyecharts import Bar, Line
from pyecharts.engine import create_default_environment

import time
import datetime
import html2text
import functools

# 用户登录装饰器


def login(func):
    @functools.wraps(func)
    def wrapper(*args, **kw):
        # 将request传入包装器内，以便查询用户登录session
        request = args[0]
        # 如果没有登录，跳转到登录页面
        if not request.session.get('uid', 0):
            request.session['referrerUrl'] = request.path
            return HttpResponseRedirect("/notebook/user/login")
        # 提取登录的用户信息
        user = User.objects.get(id=request.session['uid'])
        # 如果已经登录，执行请求的函数
        # 将登录用户信息 user作为参数添加进去
        return func(*args, **kw, user=user)
    return wrapper


# ICU质控主页
@login
def index(request, user):
    template = "icu/tools/qualityControl/index.html"
    return render(request, template, {'user': user})

# ICU质控帮助文件
@login
def index(request, user):
    template = "icu/help.html"
    return render(request, template, {'user': user})

# 编辑质控数据，显示并处理表单，如果日期未质控，转到add，如果日期已质控，转到update
@login
def edit(request, date, user):
    # GET时显示表单
    if request.method == 'GET':
        # 昨天日期
        yesterday = datetime.datetime.strptime(
            date, '%Y-%m-%d') + datetime.timedelta(days=-1)
        # 明天日期
        tomorrow = datetime.datetime.strptime(
            date, '%Y-%m-%d') + datetime.timedelta(days=1)
        # 今日质控数据，保存到qc中
        qcs = QualityControl.objects.filter(user=user, qcDate=date)
        qc = {}
        if qcs:
            qc = qcs[0]
        # 昨日质控数据，保存到qc_y中
        qcs_yesterday = QualityControl.objects.filter(
            user=user, qcDate=yesterday)
        qc_y = {}
        if qcs_yesterday:
            qc_y = qcs_yesterday[0]
        # 显示表单
        template = "icu/tools/qualityControl/edit_form.html"
        return render(request,
                      template,
                      {'user': user,
                       'qc': qc,
                       'qc_y': qc_y,
                       'qcDate': date,
                       'yesterday': yesterday,
                       'tomorrow': tomorrow,
                       })
    # 如果POST数据，处理提交的数据
    elif request.method == 'POST':
        # qc保存提交的质控数据
        qc = {}
        # 整理数据，保存在qc 中
        cleanUp(request, qc)
        qcs = QualityControl.objects.filter(user=user, qcDate=qc['qcDate'])
        if qcs:
            # 若已经存在数据，则update
            action = "update"
        else:
            # 若没有数据，则 create
            action = "create"

        # 验证数据有效性
        errors = validate(request, qc, action)
        if errors:
            template = "icu/tools/qualityControl/submitError.html"
            return render(request, template, {'errors': errors, })
        # 存储
        storeQc = store(request, user, qc, action)
        template = "icu/tools/qualityControl/submitSuccess.html"
        return render(request, template, {'qc': storeQc})


# 整理用户提交的数据，使用字典参数qc保存数据
def cleanUp(request, qc):

    # 整数类型字段填充
    intKeys = (
        'total',
        'entry',
        'outpatient',
        'shiftIn',
        'unplannedShift',
        'revert',
        'out',
        'transfer',
        'improve',
        'cure',
        'automaticDischarge',
        'death',
        'ventilation',
        'newVAP',
        'newTracheaCannula',
        'unplannedExtubation',
        'reintubation',
        'AVCatheter',
        'CRBSI',
        'urethralCatheter',
        'CAUTI',
        'septicShock',
        'bundle3',
        'bundle6',
        'antibiotic',
        'sample',
        'preventDVT',
        'newDVT',
    )
    for key in intKeys:
        qc[key] = int(request.POST.get(key) or 0)
    # 其它字段填充
    qc['qcDate'] = datetime.datetime.strptime(
        request.POST['qcDate'], '%Y-%m-%d').date()
    qc['comments'] = request.POST['comments']
    qc['doctor'] = request.POST.get('doctor', '')  # 没有填报医生数据，就填空
    qc['ip'] = request.META.get('REMOTE_ADDR', '')  # IP地址

# 审核用户提交的数据，确认数据，以保存复数准确


def validate(request, qc, action):
    errors = []

    if datetime.date.today() + datetime.timedelta(days=1) < qc['qcDate']:
        errors = ['至多允许提前一天填写质控数据', ]

    if datetime.date.today() + datetime.timedelta(days=-7) > qc['qcDate']:
        errors = ['一周前的数据不能修改', ]
    return errors


# 存储用户数据和最终结果
def store(request, user, qc, action):
    # 新建数据库数据
    if action == 'create':
        storeQc = QualityControl.objects.create(
            user=user, **qc)

    # 更新数据
    elif action == 'update':
        # up 表示update patient
        storeQc = QualityControl.objects.get(qcDate=qc['qcDate'], user=user)
        for key in qc.keys():
            setattr(storeQc, key, qc[key])
        storeQc.save()
    return storeQc


# QualityControl列表 qcList
@login
def qcList(request, page, user):
    # 格式化要查询的页码
    page = page or 1
    page = int(page)
    # 每次显示count个数据
    count = 10

    qcs_all = QualityControl.objects.filter(user=user).order_by('-qcDate')

    paginator = Paginator(qcs_all, count)

    try:
        qcs = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        qcs = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        qcs = paginator.page(paginator.num_pages)

    # 页面范围
    #page_range =paginator.page_range
    if page < 6:
        if paginator.num_pages <= 10:
            page_range = range(1, paginator.num_pages + 1)
        else:
            page_range = range(1, 11)
    elif (page >= 6) and (page <= paginator.num_pages - 5):
        page_range = range(page - 5, page + 5)
    else:
        page_range = range(paginator.num_pages - 9, paginator.num_pages + 1)

    # 计数器开始值,使用负值，以便在模板中可以使用加法tag
    start = (qcs.number - 1) * count - qcs_all.count() - 1

    # statistics 进行统计
    echarts = statistics(request)

    template = "icu/tools/qualityControl/list.html"
    return render(request,
                  template,
                  {'user': user,
                   'qcs': qcs,
                   'page_range': page_range,
                   'start': start,
                   'echarts': echarts,
                   })


#  统计数据
def statistics(request):
    from pyecharts import Bar, Line
    from pyecharts.engine import create_default_environment
    from django.db.models import Avg, Max, Min
    import numpy
    # 为渲染创建一个默认配置环境
    # create_default_environment(filet_ype)
    # file_type: 'html', 'svg', 'png', 'jpeg', 'gif' or 'pdf'

    user = User.objects.get(id=request.session['uid'])

    p = {}
    p = QualityControl.objects.filter(user=user).order_by(
        '-qcDate').values_list("qcDate", "total", "entry", "death")[0:30]
    # 如果没有返回值，则显示空值
    if not p:
        return ''
    # 矩阵转置
    p = numpy.transpose(p)

    line = Line("重症质控一览表", "近30日流量")
    line.add("科室患者数", p[0][::-1], p[1][::-1])
    line.add("收治患者数", p[0][::-1], p[2][::-1])
    line.add("死亡患者数", p[0][::-1], p[3][::-1])

    env = create_default_environment("html")
    chart1 = env.render_container_and_echarts_code(chart=line)
    echarts = [chart1, ]

    return echarts


# 提供文件下载
@login
def download(request, condition, parameter, user):
    # 以id为条件进行查询时，调用函数download_id
    if condition == 'all':
        return download_all(request, user)


def download_all(request, user):
    from openpyxl import Workbook
    # 导入模块，处理下载时中文uri乱码的问题
    from django.utils.encoding import escape_uri_path
    qcs = QualityControl.objects.filter(user=user)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '日期'
    ws['B1'] = '患者数'
    ws['C1'] = '昨日入科'
    ws['D1'] = '门急诊转入'
    ws['E1'] = '病房转入'
    ws['F1'] = '非计划转入'
    ws['G1'] = '48h重返'
    ws['H1'] = '昨总出科'
    ws['I1'] = '转出他科'
    ws['J1'] = '好转出院'
    ws['K1'] = '治愈出院'
    ws['L1'] = '自动出院'
    ws['M1'] = '死亡'
    ws['N1'] = '机械通气数'
    ws['O1'] = '新增VAP'
    ws['P1'] = '新气管插管'
    ws['Q1'] = '非计划拨管'
    ws['R1'] = '48h再插管'
    ws['S1'] = '血流导管总数'
    ws['T1'] = '新血流感染'
    ws['U1'] = '导尿管总数'
    ws['V1'] = '新尿管感染'
    ws['W1'] = '感染性休克'
    ws['X1'] = '3hBundle'
    ws['Y1'] = '6hBundle'
    ws['Z1'] = '新增抗生素'
    ws['AA1'] = '标本培养'
    ws['AB1'] = '新增DVT预防'
    ws['AC1'] = '新发DVT数'
    ws['AD1'] = '备注'

    line = 1
    for q in qcs:
        line += 1
        ws['A%d' % (line)] = (
            q.qcDate + datetime.timedelta(hours=8)).strftime("%Y-%m-%d")
        ws['B%d' % (line)] = q.total
        ws['C%d' % (line)] = q.entry
        ws['D%d' % (line)] = q.outpatient
        ws['E%d' % (line)] = q.shiftIn
        ws['F%d' % (line)] = q.unplannedShift
        ws['G%d' % (line)] = q.revert
        ws['H%d' % (line)] = q.out
        ws['I%d' % (line)] = q.transfer
        ws['J%d' % (line)] = q.improve
        ws['K%d' % (line)] = q.cure
        ws['L%d' % (line)] = q.automaticDischarge
        ws['M%d' % (line)] = q.death
        ws['N%d' % (line)] = q.ventilation
        ws['O%d' % (line)] = q.newVAP
        ws['P%d' % (line)] = q.newTracheaCannula
        ws['Q%d' % (line)] = q.unplannedExtubation
        ws['R%d' % (line)] = q.reintubation
        ws['S%d' % (line)] = q.AVCatheter
        ws['T%d' % (line)] = q.CRBSI
        ws['U%d' % (line)] = q.urethralCatheter
        ws['V%d' % (line)] = q.CAUTI
        ws['W%d' % (line)] = q.septicShock
        ws['X%d' % (line)] = q.bundle3
        ws['Y%d' % (line)] = q.bundle6
        ws['Z%d' % (line)] = q.antibiotic
        ws['AA%d' % (line)] = q.sample
        ws['AB%d' % (line)] = q.preventDVT
        ws['AC%d' % (line)] = q.newDVT
        ws['AD%d' % (line)] = html2text.html2text(q.comments).strip()

    response = HR(content_type='application/vnd.ms-excel')
    #date = time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime())
    date = time.strftime("%Y-%m-%d", time.localtime())

    filename = "%sICU质控汇总表-%s.xlsx" % (user.company, date)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(
        escape_uri_path(filename))
    wb.save(response)
    return response

# 下载报表
@login
def reportDownload(request, year, user):
    year = int(year)
    # 使用info存储整年和每月的数据
    info = []

    # 从 0 月（一整年） 到 12 月的数据
    for month in range(13):
        info.append(get_report_data(user, year, month))

    months = ['全年'] + [str(x) + "月" for x in range(1, 13)]

    from openpyxl import Workbook
    # 导入模块，处理下载时中文uri乱码的问题
    from django.utils.encoding import escape_uri_path

    wb = Workbook()
    ws = wb.active

    # 定义纵列位置对应的字段值
    rank = [chr(x) for x in range(ord('A'), ord('Z') + 1)] + \
        ["A" + chr(x) for x in range(ord('A'), ord('N') + 1)]
    # 37个指标
    values = (
        "时间",
        "收治总数",
        "收治床位日",
        "ApacheII评分人数",
        "ApacheII评分≥15分人数",
        "ApacheII≥15收治率",
        "感染性休克人数",
        "3hBundle完成数",
        "6hBundle完成数",
        "3hBundle完成率",
        "6hBundle完成率",
        "治疗性抗菌药物使用数",
        "抗生素前标本送检数",
        "标本送检率",
        "DVT预防数",
        "DVT预防率",
        "预计病死率",
        "实际病死数",
        "实际病死率",
        "标化病死率",
        "气管插管数",
        "非计划拨管数",
        "非计划拨管率",
        "48小时再插管数",
        "48h再插管率",
        "非计划术后转入数",
        "转入人数",
        "非计划转入率",
        "48小时重返数",
        "转出数",
        "48h重返率",
        "VAP发生数",
        "呼吸机使用日",
        "VAP发生率",
        "CRBSI发生数",
        "血流导管留置日",
        "CRBSI发生率",
        "CAUTI发生数",
        "留置尿管日",
        "CAUTI发生率",
    )

    for n in range(len(rank)):
        ws[rank[n] + '1'] = values[n]

    row = 1
    for data in info:
        row += 1
        # 更新属性
        keys = (
            'month',
            'entry__sum',
            'total__sum',
            'apacheII__count',
            'apacheII15__count',
            'apacheII15Rate',
            'septicShock__sum',
            'bundle3__sum',
            'bundle6__sum',
            'bundle3Rate',
            'bundle6Rate',
            'antibiotic__sum',
            'sample__sum',
            'sampleRate',
            'preventDVT__sum',
            'preventDVTRate',
            'apacheIIDeathRate',
            'death__sum',
            'realDeathRate',
            'standDeathRate',
            'newTracheaCannula__sum',
            'unplannedExtubation__sum',
            'unplannedExtubationRate',
            'reintubation__sum',
            'reintubationRate',
            'unplannedShift__sum',
            'shiftIn__sum',
            'unplannedShiftRate',
            'revert__sum',
            'transfer__sum',
            'revertRate',
            'newVAP__sum',
            'ventilation__sum',
            'VAPRate',
            'CRBSI__sum',
            'AVCatheter__sum',
            'CRBSIRate',
            'CAUTI__sum',
            'urethralCatheter__sum',
            'CAUTIRate')
        # 最后一个时间不能直接添加,需要手动加上8小时后再添加
        for n in range(len(rank)):
            ws[rank[n] + str(row)] = data[keys[n]]

    response = HR(content_type='application/vnd.ms-excel')
    filename = "%sICU质控年度报表-%s.xlsx" % (user.company, year)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(
        escape_uri_path(filename))
    wb.save(response)
    return response

# 在线报表
@login
def report(request, year, user):
    year = year or datetime.datetime.now().year
    year = int(year)

    # 使用info存储整年和每月的数据
    info = []

    # 从 0 月（一整年） 到 12 月的数据
    for month in range(13):
        info.append(get_report_data(user, year, month))

    months = ['全年'] + [str(x) + "月" for x in range(1, 13)]

    charts = []
    # 收治情况
    charts.append(Bar("%d年收治" % (year), "每月收治人数"))
    charts[-1].add("收治人数", months[1:], [monthData['entry__sum']
                                        for monthData in info[1:]])
    # 床位日
    charts.append(Bar("%d年床位日" % (year), "每月床位日"))
    charts[-1].add("床位日", months[1:], [monthData['total__sum']
                                       for monthData in info[1:]])
    # 收治率(大概率)
    charts.append(Bar("%d年" % (year), "每月统计率%"))
    charts[-1].add("ApacheII评分≥15分收治率", months,
                   [monthData['apacheII15Rate'] for monthData in info])
    charts[-1].add("预计病死率", months, [monthData['apacheIIDeathRate']
                                     for monthData in info])
    charts[-1].add("标化病死率", months, [monthData['standDeathRate']
                                     for monthData in info])
    # 收治率
    charts.append(Bar("%d年" % (year), "每月统计率%"))
    charts[-1].add("治疗性抗生素标本送检率", months, [monthData['sampleRate']
                                           for monthData in info])
    charts[-1].add("深静脉血栓预防率", months, [monthData['preventDVTRate']
                                        for monthData in info])
    charts[-1].add("3hBundle完成率", months, [monthData['bundle3Rate']
                                           for monthData in info])
    charts[-1].add("6hBundle完成率", months, [monthData['bundle6Rate']
                                           for monthData in info])

    # 收治率（小概率）
    charts.append(Bar("%d年" % (year), "每月统计率%"))
    charts[-1].add("非计划气管插管拨管率", months,
                   [monthData['unplannedExtubationRate'] for monthData in info])
    charts[-1].add("48h再插管率", months, [monthData['reintubationRate']
                                       for monthData in info])
    charts[-1].add("非计划术后入ICU率", months,
                   [monthData['unplannedShiftRate'] for monthData in info])
    charts[-1].add("48小时后重返ICU率", months, [monthData['revertRate']
                                           for monthData in info])
    # 三管监测（小概率）
    charts.append(Bar("%d年" % (year), "每月统计率‰"))
    charts[-1].add("VAP发生率", months, [monthData['VAPRate']
                                      for monthData in info])
    charts[-1].add("CRBSI发生率", months, [monthData['CRBSIRate']
                                        for monthData in info])
    charts[-1].add("CAUTI发生率", months, [monthData['CAUTIRate']
                                        for monthData in info])

    # 为渲染创建一个默认配置环境
    env = create_default_environment("html")
    echarts = []
    for chart in charts:
        echarts.append(env.render_container_and_echarts_code(chart=chart))

    template = "icu/tools/qualityControl/report.html"
    return render(
        request, template, {
            'user': user, 'info': info, 'echarts': echarts})


def get_report_data(user, year, month=0):

    # 使用Sum函数从QualityControl数据库查询结果
    qc_data_sum = (
        'total',
        'entry',
        'bundle3',
        'bundle6',
        'septicShock',
        'antibiotic',
        'sample',
        'preventDVT',
        'death',
        'unplannedExtubation',
        'newTracheaCannula',
        'reintubation',
        'unplannedShift',
        'shiftIn',
        'revert',
        'transfer',
        'newVAP',
        'ventilation',
        'CRBSI',
        'AVCatheter',
        'CAUTI',
        'urethralCatheter')
    # 使用Avg函数从QualityControl数据库查询结果
    qc_data_avg = ('total', 'entry')
    # 使用Count函数从QualityControl数据库查询结果
    qc_data_count = ('qcDate')
    # 使用Avg函数从ApacheII数据库查询结果
    apacheII_data_avg = ('deathRate')
    # 使用Count函数从ApacheII数据库查询结果
    apacheII_data_count = ('apacheII15__count')

    # 根据要求设置查询参数
    if month == 0:
        # 查询一年的数据,参数设置
        parameter_qc = {'user': user, 'qcDate__year': year}
        parameter_apacheII = {'user': user, 'scoreTime__year': year}
    else:
        # 按 年+月 查询数据,参数设置
        parameter_qc = {
            'user': user,
            'qcDate__year': year,
            'qcDate__month': month}
        parameter_apacheII = {
            'user': user,
            'scoreTime__year': year,
            'scoreTime__month': month}

    #sumParameter = (Sum(key) for key in qc_data_sum.keys())

    qcInfo = QualityControl.objects.filter(**parameter_qc).aggregate(
        *(Sum(key) for key in qc_data_sum), Avg('total'), Avg('entry'), Count('qcDate'),
    )
    # ApacheII 大于 15分人数
    apacheII15__count = ApacheII.objects.filter(
        **parameter_apacheII,
        score__gte=15).aggregate(
        Count('id'))['id__count'] or 0
    # ApacheII 总人数
    apacheII__count = ApacheII.objects.filter(
        **parameter_apacheII).aggregate(Count('id'))['id__count'] or 0
    # 预计病死率
    apacheIIDeathRate = ApacheII.objects.filter(
        **parameter_apacheII).aggregate(Avg('deathRate'))['deathRate__avg'] or 0

    # 整理数据，准备返回
    info = {}
    info['year'] = year
    info['month'] = month or '全年'
    for key in qcInfo:
        info[key] = qcInfo[key] or 0

    info['apacheII15__count'] = apacheII15__count
    info['apacheII__count'] = apacheII__count
    info['apacheIIDeathRate'] = round(apacheIIDeathRate, 2)

    if info['apacheII__count']:
        # ApacheII 大于等于 15分 收治率
        info['apacheII15Rate'] = round(
            info['apacheII15__count'] / info['apacheII__count'] * 100, 2)
    else:
        info['apacheII15Rate'] = 0

    if info['entry__sum']:
        # DVT预防率
        info['preventDVTRate'] = round(
            info['preventDVT__sum'] / info['entry__sum'] * 100, 2)
        # 实际病死率
        info['realDeathRate'] = round(
            info['death__sum'] / info['entry__sum'] * 100, 2)
    else:
        info['apacheII15Rate'] = 0
        info['preventDVTRate'] = 0
        info['realDeathRate'] = 0

    if info['apacheIIDeathRate']:
        # 标化病死率
        info['standDeathRate'] = round(
            info['realDeathRate'] / info['apacheIIDeathRate'] * 100, 2)
    else:
        info['standDeathRate'] = 0

    # Bundle完成率
    if info['septicShock__sum']:
        info['bundle3Rate'] = round(
            info['bundle3__sum'] / info['septicShock__sum'] * 100, 2)
        info['bundle6Rate'] = round(
            info['bundle6__sum'] / info['septicShock__sum'] * 100, 2)
    else:
        info['bundle3Rate'] = 0
        info['bundle6Rate'] = 0

    # 抗菌药物送检率
    if info['antibiotic__sum']:
        info['sampleRate'] = round(
            info['sample__sum'] / info['antibiotic__sum'] * 100, 2)
    else:
        info['sampleRate'] = 0

    # 气管插管相关
    if info['newTracheaCannula__sum']:
        # 非计划拨管率
        info['unplannedExtubationRate'] = round(
            info['unplannedExtubation__sum'] / info['newTracheaCannula__sum'] * 100, 2)
        info['reintubationRate'] = round(
            info['reintubation__sum'] /
            info['newTracheaCannula__sum'] *
            100,
            2)
    else:
        # 48h再插管
        info['unplannedExtubationRate'] = 0
        info['reintubationRate'] = 0

    # 转入患者
    if info['shiftIn__sum']:
        # 非计划转入患者率（特指手术患者）
        info['unplannedShiftRate'] = round(
            info['unplannedShift__sum'] / info['shiftIn__sum'] * 100, 2)
    else:
        info['unplannedShiftRate'] = 0

    # 转出患者
    if info['transfer__sum']:
        # 48小时转出后重返率
        info['revertRate'] = round(
            info['revert__sum'] / info['transfer__sum'] * 100, 2)
    else:
        info['revertRate'] = 0

    # VAP发生率 千分率
    if info['ventilation__sum']:
        info['VAPRate'] = round(
            info['newVAP__sum'] /
            info['ventilation__sum'] *
            1000,
            2)
    else:
        info['VAPRate'] = 0

    # CRBSI发生率 千分率
    if info['AVCatheter__sum']:
        info['CRBSIRate'] = round(
            info['CRBSI__sum'] /
            info['AVCatheter__sum'] *
            1000,
            2)
    else:
        info['CRBSIRate'] = 0

    # CAUTI发生率 千分率
    if info['urethralCatheter__sum']:
        info['CAUTIRate'] = round(
            info['CAUTI__sum'] /
            info['urethralCatheter__sum'] *
            1000,
            2)
    else:
        info['CAUTIRate'] = 0

    # 返回报表数据
    return info
