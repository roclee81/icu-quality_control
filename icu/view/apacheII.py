#!/usr/bin/python3
from django.shortcuts import render,render_to_response
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse as HR
from django.http import HttpResponseRedirect
from settings import BASE_DIR

from notebook.models import User
from icu.models import ApacheII

import os
import time
import datetime,pytz
from math import exp
import functools

#用户登录装饰器
def login(func):
    @functools.wraps(func)
    def wrapper(*args,**kw):
        #将request传入包装器内，以便查询用户登录session
        request = args[0]
        #如果没有登录，跳转到登录页面
        if not request.session.get('uid',0):
            request.session['referrerUrl'] = request.path
            return HttpResponseRedirect("/notebook/user/login")
        #提取登录的用户信息
        user=User.objects.get(id=request.session['uid'])
        #如果已经登录，执行请求的函数
        #将登录用户信息 user作为参数添加进去
        return func(*args,**kw,user=user)
    return wrapper


#ApacheII主页
@login
def index(request,user):

    template = "icu/tools/apacheII/index.html"
    return render(request,template,{'user':user})




#新增一个ApacheII ，处理表单
@login
def add(request,user):
    workmates = user.workmate.split(",")

    if request.method == 'GET':
        #显示表单
        template = "icu/tools/apacheII/edit_form.html"
        return render(request,template,{'user':user,'workmates':workmates})
    #如果POST数据
    elif request.method =='POST':
        p=patient = {}
        #整理数据，保存在patient/p 中
        cleanUp(request,p)
        #确认提交数据
        errors = validate(request,p,'create')
        if errors:
            template = "icu/tools/apacheII/submitError.html"
            return render(request,template,{'errors':errors,})
        #计算
        calculate(p)
        #存储
        store(request,p,'create')
        template = "icu/tools/apacheII/submitSuccess.html"
        return render(request,template,{'user':user,'id':request.session['newPatientId'],'workmates':workmates})

#修改一个已存在的ApacheII ，处理表单
@login
def update(request,id,user):
    patient=ApacheII.objects.get(user=user,id=id)
    workmates = user.workmate.split(",")

    if request.method == 'GET':
        template = "icu/tools/apacheII/edit_form.html"
        return render(request,template,{'user':user,'p':patient,'workmates':workmates})
    #如果POST数据
    elif request.method =='POST':
        p=patient = {}
        #整理数据，保存在patient/p 中
        cleanUp(request,p)
        #确认提交数据
        errors = validate(request,p,'update')
        if errors:
            template = "icu/tools/apacheII/submitError.html"
            return render(request,template,{'errors':errors,})
        #计算
        calculate(p)
        #存储
        store(request,p,'update')
        template = "icu/tools/apacheII/submitSuccess.html"
        return render(request,template,{'user':user,'id':request.session['newPatientId'],'workmates':workmates})



#整理用户提交的数据，使用字典参数p保存数据
def cleanUp(request,p):
    #非空的int值
    for intKey in ('age','inNumber','SBP','DBP','HR','R','E','V','M'):
        p[intKey] = int(request.POST.get(intKey))
    #非空的float值
    for floatKey in ('temperature','WBC','HCT','pH','FiO2','PaO2','PaCO2','Na','K','Cr',):
        p[floatKey] = float(request.POST.get(floatKey))
    #非空的普通str值
    for commonKey in ('name','bedNumber','AKI','health','doctor'):
        p[commonKey] = request.POST.get(commonKey)
    #可以为空的str值
    for nullKey in ('noSurgery','surgery'):
        p[nullKey] = request.POST.get(nullKey,'')
    #可以为0的int值
    p['id'] = int(request.POST.get('id','0'))
    p['ip'] = request.META.get('REMOTE_ADDR','') #IP地址


    #如果为非手术类疾病，不存在手术后
    if p['noSurgery']:
        p['afterSurgery'] = ''
    else:
        p['afterSurgery'] = request.POST.get('afterSurgery','')
    #存储时间
    p['scoreTime'] = datetime.datetime.strptime(request.POST['scoreTime'],'%Y-%m-%dT%H:%M')

    #计算MAP 
    p['MAP'] = float(p['SBP'])/3 + 2*float(p['DBP'])/3
    #计算AaDO2
    p['AaDO2'] = p['FiO2'] * (760 -47)*0.01 - p['PaCO2']/0.8 - p['PaO2']


#审核用户提交的数据，确认数据，以保证数据准确
def validate(request,p,action):
    user=User.objects.get(id=request.session['uid'])
    last = ApacheII.objects.filter(user=user).order_by('-id')[0:1]

    if p['noSurgery'] and p['surgery'] :
        errors =['入院主要疾病只能为一种，要么为手术类，要么为非手术类',]
        return errors

    if action =='create':
        if last and last[0].name == p['name'] :
            errors =['最近一个相同的患者姓名提交，不被允许',]
            return errors
    elif action =='update':
        up = ApacheII.objects.get(id=p['id'],user=user)
        createTime = up.createTime
        
        t = datetime.datetime.now()
        new_t=t.replace(tzinfo=(pytz.timezone('Asia/Shanghai')))
        if new_t + datetime.timedelta(days = -7) > createTime :
            errors =['一周前的数据不能修改',]
            return errors

#存储用户数据和最终结果
def store(request,p,action):
    user=User.objects.get(id=request.session['uid'])

    #如果动作为新建一个ApacheII评分表
    if action == 'create':
        #删除 P 中的 id ，id由数据库自行管理
        del(p['id'])
        newPatient = ApacheII.objects.create( user = user,**p)
        #将新提交的患者数据库id放进session中存储
        request.session['newPatientId'] = newPatient.id
    if action == 'update':
        # up 表示update patient
        up = ApacheII.objects.get(id=p['id'],user=user)


        #将患者数据库id放进session中存储
        request.session['newPatientId'] = p['id']
        #删除 P 中的 id ，id由数据库自行管理
        del(p['id'])

        #更新属性
        for key in p.keys():
            #setattr方法可以对 对象 的属性进行赋值
            setattr(up,key,p[key])
        up.save()
            

#通过提交的数据计算ApacheII值 及R值 
#参数Patient为患者信息
def calculate(p):
    scores = []
    if p['age']>=75:
        scores.append(6)
    elif p['age']>=65:
        scores.append(5)
    elif p['age']>=55:
        scores.append(3)
    elif p['age']>=45:
        scores.append(2)
    elif p['age']>=0:
        scores.append(0)
    
    if p['temperature'] >=41:
        scores.append(4)
    elif p['temperature'] >=39:
        scores.append(3)
    elif p['temperature'] >=38.5:
        scores.append(1)
    elif p['temperature'] >=36:
        scores.append(0)
    elif p['temperature'] >=34:
        scores.append(1)
    elif p['temperature'] >=32:
        scores.append(2)
    elif p['temperature'] >=30:
        scores.append(3)
    elif p['temperature'] >=0:
        scores.append(4)

    if p['MAP'] >= 160:
        scores.append(4)
    elif p['MAP'] >= 130:
        scores.append(3)
    elif p['MAP'] >= 110:
        scores.append(2)
    elif p['MAP'] >= 70:
        scores.append(0)
    elif p['MAP'] >= 50:
        scores.append(2)
    elif p['MAP'] >= 0:
        scores.append(4)


    if p['HR'] >= 180:
        scores.append(4)
    elif p['HR'] >= 140:
        scores.append(3)
    elif p['HR'] >= 110:
        scores.append(2)
    elif p['HR'] >= 70:
        scores.append(0)
    elif p['HR'] >= 55:
        scores.append(2)
    elif p['HR'] >= 40:
        scores.append(3)
    elif p['HR'] >= 0:
        scores.append(4)


    if p['R'] >= 50:
        scores.append(4)
    elif p['R'] >= 35:
        scores.append(3)
    elif p['R'] >= 25:
        scores.append(1)
    elif p['R'] >= 12:
        scores.append(0)
    elif p['R'] >= 10:
        scores.append(1)
    elif p['R'] >= 6:
        scores.append(2)
    elif p['R'] >= 0:
        scores.append(4)




    if p['WBC'] >= 40:
        scores.append(4)
    elif p['WBC'] >= 20:
        scores.append(2)
    elif p['WBC'] >= 15:
        scores.append(1)
    elif p['WBC'] >= 3:
        scores.append(0)
    elif p['WBC'] >= 1:
        scores.append(2)
    elif p['WBC'] >= 0:
        scores.append(4)


    if p['HCT'] >= 60:
        scores.append(4)
    elif p['HCT'] >= 50:
        scores.append(2)
    elif p['HCT'] >= 46:
        scores.append(1)
    elif p['HCT'] >= 30:
        scores.append(0)
    elif p['HCT'] >= 20:
        scores.append(2)
    elif p['HCT'] >= 0:
        scores.append(4)


    if p['pH'] >= 7.7:
        scores.append(4)
    elif p['pH'] >= 7.6:
        scores.append(3)
    elif p['pH'] >= 7.5:
        scores.append(1)
    elif p['pH'] >= 7.33:
        scores.append(0)
    elif p['pH'] >= 7.25:
        scores.append(2)
    elif p['pH'] >= 7.15:
        scores.append(3)
    elif p['pH'] >= 0:
        scores.append(4)




    if p['FiO2'] <50:
        if p['PaO2'] >70:
            scores.append(0)
        elif p['PaO2'] >=61:
            scores.append(1)
        elif p['PaO2'] >=55:
            scores.append(3)
        elif p['PaO2'] >=0:
            scores.append(4)
    else:
        if p['AaDO2'] >= 500:
            scores.append(4)
        elif p['AaDO2'] >= 350:
            scores.append(3)
        elif p['AaDO2'] >= 200:
            scores.append(2)
        elif p['AaDO2'] >= 0:
            scores.append(0)



    if p['Na'] >= 180:
        scores.append(4)
    elif p['Na'] >= 160:
        scores.append(3)
    elif p['Na'] >= 155:
        scores.append(2)
    elif p['Na'] >= 150:
        scores.append(1)
    elif p['Na'] >= 130:
        scores.append(0)
    elif p['Na'] >= 120:
        scores.append(2)
    elif p['Na'] >= 111:
        scores.append(3)
    elif p['Na'] >= 0:
        scores.append(4)



    if p['K'] >= 7:
        scores.append(4)
    elif p['K'] >= 6:
        scores.append(3)
    elif p['K'] >= 5.5:
        scores.append(1)
    elif p['K'] >= 3.5:
        scores.append(0)
    elif p['K'] >= 3:
        scores.append(1)
    elif p['K'] >= 2.5:
        scores.append(2)
    elif p['K'] >= 0:
        scores.append(4)


    if p['Cr'] >= 309.4:
        scores.append(4)
    elif p['Cr'] >= 176.8:
        scores.append(3)
    elif p['Cr'] >= 132.6:
        scores.append(2)
    elif p['Cr'] >= 53.04:
        scores.append(0)
    elif p['Cr'] >= 0:
        scores.append(2)

    if p['AKI'] == '是':
        scores.append(scores[-1])
    else:
        scores.append(0)

    scores.append(int(p['health'].split(":")[1]))
    scores.append(15)
    scores.append(-p['E'])
    scores.append(-p['V'])
    scores.append(-p['M'])

    #计算R值需要的Score2参数
    scores2 = []
    scores2.append(-3.517)
    if p['surgery']:
        scores2.append(float(p['surgery'].split(":")[1]))
        scores2.append(float(p['afterSurgery'].split(":")[1]))
    else:
        scores2.append(float(p['noSurgery'].split(":")[1]))


    data1 = sum(scores) * 0.146 + sum(scores2)
    R = exp(data1)/(1+exp(data1))*100

    p['score'] = sum(scores)
    p['deathRate'] = round(R,2)


#提供文件下载
@login
def download(request,condition,parameter,user):

    #以id为条件进行查询时，调用函数download_id
    if condition == 'id':
        return download_id(request,parameter)
    elif condition == 'all':
        return download_all(request)

def download_all(request):
    from openpyxl import Workbook
    #导入模块，处理下载时中文uri乱码的问题
    from django.utils.encoding import escape_uri_path
    user=User.objects.get(id=request.session['uid'])
    patients = ApacheII.objects.filter(user=user)

    wb = Workbook()
    ws = wb.active

    #定义纵列位置对应的字段值
    rank = [chr(x) for x in range(ord('A'), ord('Z') + 1)]+['AA','AB','AC','AD','AE','AF']
    values = ( '姓名', '年龄', '床号', '住院号', '体温', '收缩压', '舒张压', '平均动脉压', '心率', '呼吸', '白细胞',
             '红细胞压积', 'pH', 'FiO2', 'PaO2', 'PaCO2', 'AaDO2', 'Na+', 'K+', 'Cr', 'AKI', '既往健康', 'E', 'V',
             'M', '非手术疾病', '手术类疾病', '手术后', '评分', '死亡率', '医生', '评分日期',)
    for n in range(len(rank)):
        ws[rank[n]+'1'] = values[n]

    row = 1
    for p in patients :
        row += 1
        #更新属性
        keys= ('name','age','bedNumber','inNumber','temperature','SBP','DBP','MAP','HR','R','WBC','HCT','pH','FiO2','PaO2','PaCO2','AaDO2',\
                    'Na','K','Cr','AKI','health','E','V','M','noSurgery','surgery','afterSurgery','score','deathRate','doctor','scoreTime')
        #最后一个时间不能直接添加,需要手动加上8小时后再添加
        for n in range(len(rank)-1):
            ws[rank[n]+str(row)] = getattr(p,keys[n])

        #非模板显示时间，必须要在数据库时间的基础上手工加上8小时
        ws['AF%d'%(row)] = (p.scoreTime +datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H-%M-%S")


    response = HR(content_type='application/vnd.ms-excel')
    #date = time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime())
    date = time.strftime("%Y-%m-%d",time.localtime())

    filename="%s-ICU-ApacheII评分表汇总-%s.xlsx"%(user.company,date)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(filename))
    wb.save(response)
    return response

@login
def download_id(request,id,user):
    from openpyxl import load_workbook
    #导入模块，处理下载时中文uri乱码的问题
    from django.utils.encoding import escape_uri_path
    
    patient = ApacheII.objects.get(id=id,user=user)
    excel_path = os.path.join(BASE_DIR,"icu/view/apache-II.xlsx")

    from openpyxl.styles import NamedStyle, Font, Border, Side
    wb = load_workbook(excel_path)
    ws = wb['备份']

    ws['B1'] = user.company

    ws['C3'] = patient.name
    ws['F3'] = patient.bedNumber
    ws['H3'] = patient.inNumber
    ws['G5'] = patient.age
    ws['G6'] = patient.temperature
    ws['G7'] = patient.MAP
    ws['G8'] = patient.HR
    ws['G9'] = patient.R
    ws['G10'] = patient.HCT
    ws['G11'] = patient.WBC
    if patient.FiO2 < 50:
        ws['G12'] = patient.PaO2
    else:
        ws['G13'] = patient.AaDO2
    ws['G14'] = patient.pH
    ws['G16'] = patient.Na
    ws['G17'] = patient.K
    ws['G18'] = patient.Cr
    ws['G19'] = patient.AKI
    ws['G20'] = patient.health.split(':')[0]
    E={1:'不能睁眼',2:'刺疼睁眼',3:'呼唤睁眼',4:'自动睁眼'}
    V={1:'不能言语',2:'只能发音',3:'答非所问',4:'回答不切题',5:'回答切题'}
    M={1:'不能活动',2:'刺疼肢体伸展',3:'刺疼肢体屈曲',4:'刺疼能躲避',5:'刺疼能定位',6:'按吩咐动作'}
    ws['G21'] = E[patient.E]
    ws['G22'] = V[patient.V]
    ws['G23'] = M[patient.M]

    if patient.surgery :
        ws['G25'] = patient.surgery.split(':')[0]
        ws['G26'] = patient.afterSurgery.split(':')[0]
    else:
        ws['G24'] = patient.noSurgery.split(':')[0]

    #非模板显示时间，必须要在数据库时间的基础上手工加上8小时
    scoreTime = patient.scoreTime +datetime.timedelta(hours=8)
    ws['E29'] = scoreTime.strftime("%Y-%m-%d %H:%M")
    ws['H29'] = patient.doctor


    border = Border(left=Side(style='medium',color='FF000000'),right=Side(style='medium',color='FF000000'),top=Side(style='medium',color='FF000000'),bottom=Side(style='medium'))
    ranges = ws['B4':'H27']

    #修复无边框的Bug
    for row in ranges:
        for cell in row:
            cell.border = border

    response = HR(content_type='application/vnd.ms-excel')
    date = time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime())

    filename="%s-ApacheII-%s.xlsx"%(patient,date)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(filename))
    wb.save(response)
    return response







#ApacheII查看
@login
def retrieve(request,id,user):
    patient=ApacheII.objects.get(user=user,id=id)

    template = "icu/tools/apacheII/retrieve.html"
    return render(request,template,{'user':user,'p':patient})

#ApacheII列表
@login
def apacheIIList(request,page,user):

    #格式化要查询的页码
    page = page or 1
    page = int(page)


    #每次显示count个数据
    count =10

    #查询到所有的患者
    patients_all = ApacheII.objects.filter(user=user).order_by('-id')

    #定义患者的分页器，每页显示count个数据
    paginator = Paginator(patients_all,count)

    try:
        #返回指分页器定页的结果
        patients = paginator.page(page) # patients为Page对象！
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        patients = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        patients = paginator.page(paginator.num_pages)

    #页面范围
    #page_range =paginator.page_range
    if page < 6 :
        if paginator.num_pages <= 10:
            page_range = range(1,paginator.num_pages + 1)
        else:
            page_range = range(1,11)
    elif (page >=6) and (page <= paginator.num_pages - 5):
        page_range = range(page - 5, page + 5)
    else:
        page_range = range(paginator.num_pages -9, paginator.num_pages +1)



    #计数器开始值,使用负值，以便在模板中可以使用加法tag
    #用于给返回的患者填序号
    start = (patients.number-1) * count  - patients_all.count() - 1 

    # pyecharts 返回统计图表
    echarts = statistics(request)

    template = "icu/tools/apacheII/list.html"
    return render(request,template,{'user':user,'patients':patients,'page_range':page_range,'start':start,'echarts':echarts})
   




#ApacheII  统计数据
def statistics(request):
    from pyecharts import Bar, Line
    from pyecharts.engine import create_default_environment
    from django.db.models import Avg, Max, Min

    user=User.objects.get(id=request.session['uid'])
    #先查询有几名医生提交了数据
    a =ApacheII.objects.filter(user=user).values('doctor').distinct().order_by('doctor')

    #医生列表
    doctor =[]
    #本月收治患者数
    monthData =[]
    #每人收治患者总数
    total =[]
    #ApacheII平均值
    ave =[]
    for aa in a:
        doctor.append( aa['doctor'] )
        #查询总数据，添加到列表中
        total.append(len(ApacheII.objects.filter(user=user,doctor=aa['doctor'])))
        #查询当前年，当前月份数据，添加到列表中
        monthData.append(len(ApacheII.objects.filter(user=user,doctor=aa['doctor'],createTime__year=datetime.datetime.now().year,createTime__month=datetime.datetime.now().month)))
        ave.append(ApacheII.objects.filter(user=user,doctor=aa['doctor']).aggregate(Avg('score'))['score__avg'])


    bar1 = Bar("本月收治", "%d月数据"%(datetime.datetime.now().month))
    bar1.add("患者数", doctor,monthData)

    bar2 = Bar("总收治患者数目", "数量对比")
    bar2.add("患者数", doctor,total)

    bar3 = Bar("ApacheII平均值", "质量对比")
    bar3.add("ApacheII平均分", doctor, ave)

    # 为渲染创建一个默认配置环境
    # create_default_environment(filet_ype)
    # file_type: 'html', 'svg', 'png', 'jpeg', 'gif' or 'pdf'
    env = create_default_environment("html")
    chart1 = env.render_container_and_echarts_code(chart=bar1)
    chart2 = env.render_container_and_echarts_code(chart=bar2)
    chart3 = env.render_container_and_echarts_code(chart=bar3)
    echarts = [chart1,chart2,chart3]

    return echarts

